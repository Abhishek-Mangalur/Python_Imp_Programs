from openpyxl import Workbook, load_workbook
import time

book = Workbook()
sheet = book.active
sheet['A1'] = 56
sheet['A2'] = 43

now = time.strftime("%x")
sheet['A3'] = now
book.save("sample2.xlsx")

wb = load_workbook("sample2.xlsx")
sheet = wb.worksheets[0]
sheet2 = wb.create_sheet('demo')
name = ["A","B","C","D","E"]
salary = ["$100","$200","$300","$400","$500"]

sheet2.cell(row = 1, column = 1).value = "Name"
sheet2.cell(row = 1, column = 2).value = "Salary"
j = 2

for i in range(0,5):
    sheet2.cell(row = j, column = 1).value = name[i]
    sheet2.cell(row = j, column = 2).value = salary[i]
    j += 1
    
print(wb.active)
sheet.cell(row = 6, column = 9).value = "Raj"
sheet['F10'] = 100
print(wb.active)
print(wb.sheetnames)
wb.save("sample2.xlsx")

print("Reading row1", sheet['A1'].value)
print("Reading row3", sheet['A3'].value)
cell = sheet['A4':'B10']

for c1, c2 in cell:
    print(c1.value, c2.value)